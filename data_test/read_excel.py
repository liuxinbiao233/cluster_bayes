import openpyxl

wb=openpyxl.load_workbook(r'C:\Users\DaBiao\Downloads\vichel\data.xlsx')
# Getting sheets from the workbook
# print(wb.sheetnames)
# for sheet in wb:
#     print(sheet.title)
ws=wb.active
# print(ws)
# print(ws['A1'])
# print(ws['A1'].value)

# c=ws['A3']
# print('Cell {} is {}'.format(c.coordinate,c.value))
# print(ws.cell(row=3,column=2))
# print(ws.cell(row=3,column=2).value)
# for i in range(1,8,2):
#     print(i,ws.cell(row=i,column=2).value)





# getting rows and columns from the sheets

col_range=ws['A:B']
row_range=ws[20:25]
# for col in col_range:
#     for cell in col:
#         print(cell.value)

# for row in row_range:
#     for cell in row:
#         print(cell.value)

# for row in ws.iter_rows(min_row=1,max_row=20,max_col=2):
#     for cell in row:
#         print(cell.value)

print('{} * {}'.format(ws.max_row,ws.max_column))

from openpyxl.utils import get_column_letter,column_index_from_string
print(get_column_letter(2),get_column_letter(78),get_column_letter(699))
print(column_index_from_string('APH'))

for i in range(10):
    print(i)
